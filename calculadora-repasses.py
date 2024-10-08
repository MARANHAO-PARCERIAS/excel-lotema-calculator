import openpyxl
from rich import print
import datetime
import logging
import locale

# Configuração do logging
logging.basicConfig(filename='calculo_repasses.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Configurar a localização para usar vírgula como separador decimal
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

def calcular_valores_em_reais(valor_base, modalidades):
    """Calcula o valor em reais para cada modalidade."""
    valores_em_reais = {}
    total = 0
    for modalidade, taxa in modalidades.items():
        valor_em_reais = taxa * valor_base
        valores_em_reais[modalidade] = round(valor_em_reais, 2)
        total += valor_em_reais
    valores_em_reais["Total"] = round(total, 2)
    return valores_em_reais

def salvar_resultado(nome_arquivo, tipo_calculo, modalidades):
    """Salva os resultados em um arquivo XLSX com data/hora e colunas ajustadas."""
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Definindo largura das colunas
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 20

        sheet['A1'] = tipo_calculo
        sheet['B1'] = "Modalidade"
        sheet['C1'] = "Valor em Reais"

        linha = 2
        for modalidade, valor in modalidades.items():
            sheet.cell(row=linha, column=1).value = tipo_calculo
            sheet.cell(row=linha, column=2).value = modalidade
            sheet.cell(row=linha, column=3).value = valor
            linha += 1

        workbook.save(nome_arquivo)
        logging.info(f"Arquivo salvo com sucesso: {nome_arquivo}")
    except Exception as e:
        logging.error(f"Erro ao salvar arquivo: {e}")

def exibir_cabecalho():
    """Exibe o cabeçalho do programa com formatação rich."""
    print("╔══════════════════════════════════════╗")
    print("║        [green]Calculadora de Repasses[/green]       ║")
    print("║    [green]LOTEMA -  Loteria do Estado MA[/green]    ║")
    print("╚══════════════════════════════════════╝")

def main():
    """Função principal que gerencia o programa."""
    exibir_cabecalho()

    try:
        # Escolha do tipo de cálculo
        print("\nEscolha o tipo de cálculo:")
        print("[bold blue]1[/bold blue]  - Prognóstico Numérico")
        print("[bold blue]2[/bold blue]  - Quota Fixa")
        print("[bold blue]3[/bold blue]  - Instantânea")

        tipo_calculo = input("Digite o número da opção desejada: ")

        # Validação da escolha do tipo de cálculo
        while tipo_calculo not in ["1", "2", "3"]:
            print("[bold red] Opção inválida. Digite 1, 2 ou 3.[/]")
            tipo_calculo = input("Digite o número da opção desejada: ")

        # Definição das modalidades e taxas
        if tipo_calculo == "1":
            tipo_calculo_texto = "PROGNOSTICO NUMERICO"
            modalidades = {
                "Educação": 0.05,
                "MAPA": 0.015,
                "P.P. Prev. e Comb. a Desas.": 0.015,
                "Seg. Social": 0.015,
                "P.P. Infân. e Juven.": 0.015
            }
        elif tipo_calculo == "2":
            tipo_calculo_texto = "QUOTA FIXA"
            modalidades = {
                "Educação": 0.02,
                "Seg. Social": 0.015,
                "Times": 0.015
            }
        elif tipo_calculo == "3":
            tipo_calculo_texto = "INSTANTÂNEA"
            modalidades = {
                "Educação": 0.05,
                "MAPA": 0.015,
                "P.P. Prev. e Comb. a Desas.": 0.015,
                "Seg. Social": 0.015,
                "P.P. Infân. e Juven.": 0.015,
                
                
            }

        # Solicitação do valor do GGR com tentativa de entrada válida
        while True:
            valor_base = input(f"\nDigite o valor do GGR de {tipo_calculo_texto}: ")

            # Ajustando o formato do valor de entrada
            if valor_base.count(',') == 1 and valor_base.count('.') <= 1: 
                valor_base = valor_base.replace(',', '.')  
            else:
                valor_base = valor_base.replace(',', '')  

            try:
                valor_base = round(float(valor_base), 2)
                break  # Sai do loop se o valor for válido
            except ValueError as e:
                mensagem_erro = "Erro: Valor numérico inválido. Por favor, insira um valor numérico válido."
                logging.error(f"Erro ao converter o valor do GGR. Entrada inválida: {valor_base}")
                print(f"[bold red]{mensagem_erro}[/]")
                # Continua no loop até que o valor seja válido

        # Calculando os valores em reais
        valores_em_reais = calcular_valores_em_reais(valor_base, modalidades)

        # Exibindo os resultados
        print("\nRESULTADOS:")
        print(f"{tipo_calculo_texto},Modalidade,Valor em Reais")
        for modalidade, valor in valores_em_reais.items():
            print(f"{tipo_calculo_texto},{modalidade},{valor:n}")  # Exibe com vírgula

        # Obtém a data e hora atuais
        agora = datetime.datetime.now()
        data_hora = agora.strftime("%d-%m-%Y_%H-%M-%S")

        # Salvando os resultados
        nome_arquivo = f"resultados_{tipo_calculo_texto.lower().replace(' ', '_')}_{data_hora}.xlsx"
        salvar_resultado(nome_arquivo, tipo_calculo_texto, valores_em_reais)

        print(f"\n[bold green] Os resultados foram salvos em '{nome_arquivo}'.[/]")
    except Exception as e:
        logging.error(f"Erro durante a execução do programa: {e}")
        print(f"[bold red]Erro: {e}[/]")

    input("\nPressione Enter para sair...")

if __name__ == "__main__":
    main()
